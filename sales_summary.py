import glob
import os
import re
import tkinter as tk
from tkinter import CENTER, E, N, W, ttk
from tkinter import filedialog as fd
from datetime import datetime
from datetime import date

# Check if all the non-default modules are downloaded. If not, download them and try again
try:
    from openpyxl import load_workbook
    import pyexcel as p
    from xhtml2pdf import pisa             
    from pydrive.drive import GoogleDrive
    from pydrive.auth import GoogleAuth
except ImportError:
  print("Trying to Install required module: openpyxl\n")
  os.system('python -m pip install openpyxl')
  print("Trying to Install required module: pyexcel\n")
  os.system('python -m pip install pyexcel')
  print("Trying to Install required module: xhtml2pdf\n")
  os.system('python -m pip install xhtml2pdf')
  print("Trying to Install required module: pydrive\n")
  os.system('python -m pip install pydrive')

from openpyxl import load_workbook
import pyexcel as p
from xhtml2pdf import pisa             
from pydrive.drive import GoogleDrive
from pydrive.auth import GoogleAuth         

cash_log = {}
root = tk.Tk()
openButton = ttk.Button()
yesButton = ttk.Button()
noButton = ttk.Button()
parsingProgressBar = ttk.Progressbar()
textProgress = tk.Label()
welcomeMessage = tk.Label()
totalFiles = 0
progressAmount = 0
folderPath = ""

def main():
    # Create results folder if it doesn't exist
    path = "./results/"
    isExist = os.path.exists(path)
    if not isExist:
        os.makedirs(path)

    # Create unique name for folder where results will be stored
    now = datetime.now()
    currentTime = now.strftime("%H.%M.%S")
    today = date.today()
    currentDay = today.strftime("%m-%d-%Y")
    global folderPath
    folderPath = currentDay + "_" + currentTime
    
    # make the new folder
    newPath = path + folderPath
    isExist = os.path.exists(newPath)
    if not isExist:
        os.makedirs(newPath)

    openFrontEnd()
    

def convertHTMLToPDF():
    os.chdir("./results/" + folderPath)
    # Use the HTML created and a new PDF file with pisa, which creates a PDF
    for data_file in glob.glob("*.html"):
        new_data_file = re.split(".html", data_file)
        new_data_file = new_data_file[0]
        output_filename = new_data_file + ".pdf"
        result_file = open(output_filename, "w+b")
        source_html = open(data_file)

        pisa.CreatePDF(source_html, dest=result_file)           
        result_file.close()              
        source_html.close()
        
        # Delete html files, it is not longer needed
        os.remove(data_file)

def createCashLog():

    total_cash_added = 0.0
    global cash_log
    path = "./results/" + folderPath + "/" + folderPath + "_cash_log.txt"

    # Store every cash log value in a text file and add them together
    with open(path, 'w') as f:
        for item in sorted(cash_log):
            value = round(cash_log[item], 2)
            f.write(item + ": " + str(value) + "\n")
            total_cash_added = total_cash_added + value
        total_cash_added = "{:.2f}".format(total_cash_added)
        f.write("TOTAL CASH: " + str(total_cash_added))

# TODO: create progress bar that does not freeze for uploading files to drive, look into threading
def uploadToDrive():
    os.chdir("../")
    path = "./results/" + folderPath 
    print("Connecting to Google Drive...")

    # Connect to the Google Drive API and authenticate
    # TODO: Store authentication information so you don't have to login every time
    try:
        gauth = GoogleAuth()
        gauth.LocalWebserverAuth()       
        drive = GoogleDrive(gauth)
    except Exception as e:
        print("Failed! Error: ")
        print(e)
    print("Connection Success!")
    print("Uploading files to Google Drive...")

    ## Upload every file that has been created from the passed excels (ignore the cash log)
    try:
        for x in os.listdir(path):
            if "cash_log" not in x:
                f = drive.CreateFile({'title': x, 'parents':[{'id': "1FDo0tvuv0onLAfF2QItgMG1lC6_MpoN-"}]})
                f.SetContentFile(os.path.join(path, x))
                f.Upload()
                f = None
            print(str(x) + " uploaded successfully")
        print("Uploading files success!")
    except Exception as e:
        print("Failed! Error: ")
        print(e)
    
    exitMessage()

def parseData(file):
    value_food = 0.0
    value_wine = 0.0
    value_beer = 0.0
    value_liquor = 0.0
    sales_tax = 0.0
    total_sales = 0.0
    total_credit_cards = 0.0
    total_cash_deposits = 0.0
    date = ""
    print("Attempting to parse through " + str(file) + "...")
    try:
        # Load the excel sheet and extract data from specific rows
        wb = load_workbook(file)

        ws = wb["Summary"]
        all_rows = list(ws.rows)

        # grab date first
        row_2 = all_rows[1]
        if row_2[0].value is not None:
            date = row_2[0].value
            # if there is a date range, grab the first date if they are the same values. Otherwise use both
            if ' ' in str(date):
                split_date = str.split(str(date), ' - ')
                # check if the dates are the same
                if split_date[0] == split_date[1]:
                    date = split_date[0]
                else:
                    date = split_date[0] + "-" + split_date[1]
        
        # make sure day was not closed
        row_5 = all_rows[4]
        if row_5[1].value != "$0.00":
            try:
                # Find food row
                food_index = findIndex(all_rows, "Food")
                row_food = all_rows[food_index]
                
                if row_food[6].value is not None:
                    value_food = row_food[6].value
                    value_food = round(value_food, 2)

                    # Add any "No Category" values, which is the next row down
                    row_no_category = all_rows[food_index + 1]
                    if row_no_category[6].value is not None:
                        value_no_category = row_no_category[6].value
                        value_food = float(value_food) + float(value_no_category)
                value_food = "{:.2f}".format(value_food)

                # Find wine row
                wine_index = findIndex(all_rows, "Wine")
                row_wine = all_rows[wine_index]
                if row_wine[6].value is not None:
                    value_wine = row_wine[6].value
                    value_wine = round(value_wine, 2)
                value_wine = "{:.2f}".format(float(value_wine))

                # Find beer row
                # TODO: Ask if i should add draft beer and bottled beer if it ever exists?
                beer_index = findIndex(all_rows, "Beer")
                row_beer = all_rows[beer_index]
                if row_beer[6].value is not None:
                    value_beer = row_beer[6].value
                    value_beer = round(value_beer, 2)
                value_beer = "{:.2f}".format(value_beer)

                # Find Liquor Row
                liquor_index = findIndex(all_rows, "Liquor")
                row_liquor = all_rows[liquor_index]
                if row_liquor[6].value is not None:
                    value_liquor = row_liquor[6].value
                    value_liquor = round(value_liquor, 2)
                value_liquor = "{:.2f}".format(value_liquor)
                    
                # Find Tax    
                row_5 = all_rows[4]
                if row_5[2].value is not None:
                    sales_tax = row_5[2].value
                    if "$" in str(sales_tax):
                        sales_tax = sales_tax.replace("$", "")
                    # remove comma if necessary
                    if "," in str(sales_tax):
                        sales_tax = sales_tax.replace(",", "")
                    sales_tax = round(float(sales_tax), 2)
                sales_tax = "{:.2f}".format(sales_tax)

                # Calculate total sales
                total_sales = float(value_food) + float(value_beer) + float(value_liquor) + float(value_wine) + float(sales_tax)
                total_sales = "{:.2f}".format(float(total_sales))

                # Calculate credit card deposits
                row_12 = all_rows[11]
                if row_12[8].value is not None:
                    total_credit_cards = row_12[8].value
                    total_credit_cards = round(total_credit_cards, 2)
                total_credit_cards = "{:.2f}".format(total_credit_cards)

                # Calculate cash deposits
                total_cash_deposits = float(total_sales) - float(total_credit_cards)
                total_cash_deposits = round(total_cash_deposits, 2)
                total_cash_deposits = "{:.2f}".format(total_cash_deposits)

                # Save cash log in dictionary
                date = checkDate(date)
                cash_log[date] = float(total_cash_deposits)
                print("Finished parsing through " + str(file))
            except Exception as e:
                print("ERROR ON DATE: " + str(date) + "\n")
                print(e)
        else:
            cash_log[date] = float(0.00)

        # Use the html template and pass the corresponding variables
        index = open("./templates/template.html").read().format(var_date=str(date),
        var_food = str(value_food), var_wine=str(value_wine), var_beer=str(value_beer), var_liquor=str(value_liquor),
        var_tax = str(sales_tax), var_total_sales=str(total_sales), var_total_credit=str(total_credit_cards), 
        var_cash = str(total_cash_deposits))
        index = index.replace("styles.css", "../../templates/styles.css")

        # save the html with the replaced variables
        date_saved = date.replace("/", "-")
        file_name = str(date_saved) + "_report.html"
        newPath = "./results/" + folderPath + "/" + file_name
        with open(newPath, 'w') as f:
            f.write(index)

        # Update front-end widgets
        root.update_idletasks()
        global parsingProgressBar
        parsingProgressBar['value'] += progressAmount 

        global textProgress
        percentage = round(parsingProgressBar['value'], 2)
        textProgress['text'] = percentage,'%'
    
    except Exception as e:
       print("Failed to create file")
       print(e)

# tidy up dates for cash log
def checkDate(date):
    new_date = date
    if "/1/" in date:
        new_date = date.replace("/1/", "/01/")
    if "/2/" in date:
        new_date = date.replace("/2/", "/02/") 
    if "/3/" in date:
        new_date = date.replace("/3/", "/03/")          
    if "/4/" in date:
        new_date = date.replace("/4/", "/04/")
    if "/5/" in date:
        new_date = date.replace("/5/", "/05/")
    if "/6/" in date:
        new_date = date.replace("/6/", "/06/")
    if "/7/" in date:
        new_date = date.replace("/7/", "/07/")
    if "/8/" in date:
        new_date = date.replace("/8/", "/08/")
    if "/9/" in date:
        new_date = date.replace("/9/", "/09/")
    return new_date

# Find the row of a given string in an excel sheet
def findIndex(rowList, stringVal):
    i = 0
    while i in range(len(rowList)):
        row_i = rowList[i]
        val = row_i[1].value
        if re.search(r"^" + stringVal, str(val)):
            return i
        i = i + 1
    return -1

def openFrontEnd():
    # Create basic front-end page
    root.title('Excel Information Extracter')
    root.geometry('750x500')
    global welcomeMessage
    welcomeMessage = tk.Label(root, text="Welcome to Excel Information Extracter! Press the button below to select excel files that you want to be parsed through. Please note this has only been configured to work with Little Mexico excel sheets.",
    wraplength=500, font=("Arial", 16))
    welcomeMessage.place(relx=0.5, rely=0.10, anchor=N)
    s = ttk.Style()
    s.configure(".", font=("Arial", 12))
    global openButton
    openButton = ttk.Button(root, text='Open Files', command=parseFiles)
    openButton.place(relx=0.5, rely=0.5, anchor=CENTER, height=50, width=100)
    root.mainloop()

def parseFiles():

    # Ask for excel files
    filetypes = (('Excel files', '.xlsx .xls'), ('All files', '*.*'))
    filenames = fd.askopenfilenames(title='Open files', initialdir='/', filetypes=filetypes)
    
    # Files are selected so button no longer necessary
    openButton.destroy()

    global totalFiles
    totalFiles = len(filenames)

    if totalFiles == 0:
        exit()

    # Update widgets
    global progressAmount
    progressAmount = 100 / totalFiles
    global parsingProgressBar
    parsingProgressBar = ttk.Progressbar(root, orient='horizontal', length=100, mode='determinate')
    parsingProgressBar.place(relx=0.5, rely=0.5, anchor=CENTER, height=50, width=100)
    parsingProgressBar['value'] = 0
    global textProgress
    textProgress = tk.Label(root, text = "0%", font=("Arial", 10))
    textProgress.place(relx=0.58, rely=0.5, anchor=W)

    # Convert old excel sheets to new ones
    for data_file in filenames:
        if data_file.endswith(".xls"):
            new_data_file = re.split(".xls", data_file)
            new_data_file = new_data_file[0]
            p.save_book_as(file_name=data_file, dest_file_name=new_data_file + ".xlsx")
            os.remove(data_file)
            parseData(new_data_file + ".xlsx")
        else:
            parseData(data_file)

    # Create cash log after parsing is finished
    createCashLog()

    # Create PDFs after cash log and parsing is finished
    convertHTMLToPDF()

    parsingProgressBar.destroy()
    textProgress.destroy()

    os.chdir("../")

    welcomeMessage.config(text = "Parsing complete. Do you want to upload to Google Drive?")

    global yesButton
    yesButton = ttk.Button(root, text='Yes', command=uploadToDrive)
    yesButton.place(relx=0.45, rely=0.5, anchor=E, height=50, width=100)

    global noButton
    noButton = ttk.Button(root, text='No', command=exit)
    noButton.place(relx=0.55, rely=0.5, anchor=W, height=50, width=100)
    
# After uploading to Google Drive is complete, let the user know everything is finished
def exitMessage():
    yesButton.destroy()
    noButton.destroy()

    global welcomeMessage
    welcomeMessage.config(text="Uploading complete. Please close this window.")

if __name__ == '__main__':
    main()